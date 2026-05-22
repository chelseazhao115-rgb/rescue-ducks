import openpyxl
import json
import re
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

# ── 1. READ DATA ──────────────────────────────────────────────
wb = openpyxl.load_workbook('renew_paraphrasing excel.xlsx')
ws = wb.active

raw_groups = []
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    seq, kw, ch, reps = row
    if kw is None or str(kw).strip() == '':
        continue
    raw_groups.append({
        'seq': str(seq).strip() if seq else '',
        'keyword': str(kw).strip(),
        'chinese': str(ch).strip() if ch else '',
        'replacements': str(reps).strip() if reps else ''
    })

print(f'Total raw groups: {len(raw_groups)}')

# ── 2. CLEANING ───────────────────────────────────────────────

# Fix known typos — comprehensive scan
TYPO_FIXES = {
    'nationa': 'national',
    'evitable': 'inevitable',
    'onlye': 'only',
    'military.': 'military',
    'witness.': 'witness',
    'Underlie': 'underlie',
    'Imperative': 'imperative',
    'Coin': 'coin',
    'fist': 'first',
    'adi.': 'adj.',
}

def clean_word(w):
    w = w.strip().strip(',').strip('.').strip('*').strip()
    if w in TYPO_FIXES:
        w = TYPO_FIXES[w]
    # Fix merged words (lowercase followed by lowercase without space)
    w = re.sub(r'([a-z])([A-Z])', r'\1, \2', w)  # camelCase split
    # Common merged patterns from the data
    merged_fixes = {
        'nonethelessnevertheless': 'nonetheless, nevertheless',
        'notwithstandingthough': 'notwithstanding, though',
        'besides,on': 'besides, on',
        'both...and': 'both...and',
        'not only...butalso': 'not only...but also',
        'neither...nor': 'neither...nor',
        'either...or': 'either...or',
        'as well as': 'as well as',
        'other than': 'other than',
        'in addition': 'in addition',
        'on the one hand...on the other hand': 'on the one hand...on the other hand',
    }
    for merged, fixed in merged_fixes.items():
        w = w.replace(merged, fixed)
    return w.strip().strip(',').strip()

def parse_replacements(rep_str):
    """Split replacement string into individual words/phrases."""
    if not rep_str or rep_str == '':
        return []

    # Pre-process: fix common data issues
    rep_str = rep_str.strip()

    # Split by comma, being careful with phrases
    parts = re.split(r',\s*', rep_str)
    words = []
    for p in parts:
        w = clean_word(p)
        if w and w != '' and w != ' ':
            words.append(w)

    return words

cleaned_groups = []
seen_keys = set()

for g in raw_groups:
    kw = clean_word(g['keyword'])
    ch = g['chinese']
    # Fix typos in Chinese definitions too
    ch = ch.replace('adi.', 'adj.')
    reps = parse_replacements(g['replacements'])

    if not reps:
        continue

    # Dedup key
    dedup_key = f"{kw.lower()}|{ch}"
    if dedup_key in seen_keys:
        continue
    seen_keys.add(dedup_key)

    # Dedup within replacements
    unique_reps = []
    seen_rep = set()
    for r in reps:
        r_lower = r.lower()
        if r_lower not in seen_rep and r_lower != kw.lower():
            unique_reps.append(r)
            seen_rep.add(r_lower)

    cleaned_groups.append({
        'seq': g['seq'],
        'keyword': kw,
        'chinese': ch,
        'replacements': unique_reps
    })

print(f'After cleaning: {len(cleaned_groups)} groups')

# ── 3. CLASSIFICATION ─────────────────────────────────────────

# Logic relation keywords for detection
LOGIC_PATTERNS_CN = [
    '指代考点', '并列考点', '转折考点', '因果考点', '比较考点',
    '举例考点', '让步考点', '条件考点', '否定考点', '并列',
    '转折', '因果', '指代', '比较', '举例', '让步', '条件',
    '否定', '考点'
]

LOGIC_KW_SET = {
    'that', 'this', 'it', 'they', 'those', 'these', 'such',
    'and', 'or', 'but', 'yet', 'however', 'whereas', 'nonetheless',
    'nevertheless', 'although', 'though', 'instead', 'rather than',
    'thanks to', 'because', 'since', 'for', 'as', 'therefore', 'hence',
    'thus', 'so', 'thereby', 'whereby',
    'if', 'unless', 'whether', 'despite', 'in spite of', 'while',
    'not only', 'both', 'neither', 'either',
    'in order to', 'so that', 'such that', 'given that', 'provided',
    'owing to', 'due to', 'because of', 'on account of',
    'as a result of', 'leading to',
    'in addition', 'besides', 'moreover', 'furthermore',
    'likewise', 'similarly', 'conversely', 'in contrast',
    'for example', 'for instance', 'such as',
    'stem from', 'derive',
    'as well as', 'other than',
    'on the one hand', 'on the other hand',
}

def classify_group(g):
    kw = g['keyword'].lower().strip()
    ch = g['chinese']
    reps = g['replacements']

    # ── Logic relations ──
    # Check Chinese definition for logic markers
    for pat in LOGIC_PATTERNS_CN:
        if pat in ch:
            return 'logic_relation'

    # Check keyword against logic set (EXACT match only — startswith is too broad for short words like "so"/"or"/"as")
    for lk in LOGIC_KW_SET:
        if kw == lk:
            return 'logic_relation'
        # For multi-word phrases, also check if kw is a sub-phrase
        if ' ' in lk and kw in lk.split(' '):
            return 'logic_relation'

    # Check if keyword is a conjunction/preposition connector
    connector_set = {
        'therefore', 'hence', 'thus', 'however', 'nevertheless', 'nonetheless',
        'moreover', 'furthermore', 'besides', 'whereas', 'although',
        'because', 'since', 'unless', 'despite', 'whereby', 'thereby',
        'consequently', 'accordingly',
    }
    if kw in connector_set:
        return 'logic_relation'

    # ── Theme clusters ──
    theme_map = {
        'nature_environment': {
            'cn': ['自然', '环境', '生态', '动物', '植物', '海洋', '河流', '气候', '地理', '森林', '物种', '栖息', '天气', '地球', '水', '土', '火', '空气', '太阳', '月', '星', '矿物', '地震', '火山', '洪水', '干旱', '风暴', '灾难'],
            'en': ['earth', 'ocean', 'marine', 'forest', 'climate', 'weather', 'environment', 'soil', 'species', 'habitat', 'ecosystem', 'wildlife', 'plant', 'animal', 'river', 'mountain', 'water', 'air', 'fire', 'sun', 'moon', 'star', 'flood', 'drought', 'storm', 'earthquake', 'natural', 'wild', 'grass', 'tree', 'flower', 'bird', 'fish', 'insect', 'coral', 'reef', 'island', 'coast', 'tide', 'wave', 'wind', 'rain', 'snow', 'ice', 'glacier', 'volcano', 'mineral', 'rock', 'stone', 'land', 'landscape', 'scenery']
        },
        'time_change': {
            'cn': ['时间', '变化', '发展', '过程', '阶段', '频率', '古代', '现代', '当代', '永久', '暂时', '长期', '短期', '延迟', '提前', '快', '慢', '古', '新', '旧', '年代', '世纪', '年度', '季节'],
            'en': ['ancient', 'modern', 'contemporary', 'current', 'future', 'past', 'present', 'chronic', 'temporary', 'permanent', 'eternal', 'transient', 'ephemeral', 'duration', 'period', 'era', 'century', 'decade', 'annual', 'yearly', 'monthly', 'weekly', 'daily', 'rapid', 'slow', 'quick', 'fast', 'delayed', 'postponed', 'prolonged', 'age', 'old', 'new', 'early', 'late', 'timely', 'seasonal']
        },
        'size_quantity': {
            'cn': ['数量', '测量', '大小', '比例', '统计', '计算', '巨大', '微小', '大', '小', '多', '少', '尺寸', '规模', '范围', '重量', '长度', '宽度', '高度', '深度', '体积'],
            'en': ['enormous', 'massive', 'tiny', 'huge', 'large', 'small', 'vast', 'immense', 'minute', 'gigantic', 'substantial', 'considerable', 'moderate', 'slight', 'significant', 'negligible', 'quantity', 'amount', 'size', 'scale', 'proportion', 'ratio', 'measure', 'weigh', 'length', 'width', 'height', 'depth', 'volume']
        },
        'health_medicine': {
            'cn': ['健康', '疾病', '医疗', '治疗', '药物', '免疫', '病毒', '感染', '诊断', '症状', '手术', '医院', '病人', '医生', '疼痛', '死亡', '生命', '康复', '愈合', '细胞', '基因', '遗传', '生理', '心理', '精神'],
            'en': ['health', 'disease', 'illness', 'medicine', 'medical', 'treatment', 'drug', 'immune', 'immunity', 'virus', 'infection', 'diagnosis', 'symptom', 'surgery', 'hospital', 'patient', 'doctor', 'pain', 'death', 'life', 'cure', 'therapy', 'heal', 'cell', 'gene', 'genetic', 'biological', 'mental', 'physical', 'body']
        },
        'society_culture': {
            'cn': ['社会', '文化', '传统', '家庭', '教育', '语言', '政治', '法律', '经济', '政府', '国家', '城市', '农村', '人口', '移民', '阶级', '性别', '种族', '宗教', '历史', '考古', '艺术', '音乐', '建筑', '文学'],
            'en': ['social', 'culture', 'tradition', 'family', 'education', 'language', 'political', 'legal', 'economic', 'government', 'country', 'city', 'urban', 'rural', 'population', 'migration', 'immigrant', 'class', 'gender', 'race', 'religion', 'history', 'archaeology', 'art', 'music', 'architecture', 'literature', 'community', 'society']
        },
        'science_technology': {
            'cn': ['科学', '技术', '实验', '研究', '数据', '数字', '计算', '工程', '机械', '电子', '化学', '物理', '生物', '数学', '信息', '网络', '电脑', '程序', '算法', '卫星', '太空', '能源', '燃料', '交通', '运输'],
            'en': ['science', 'technology', 'experiment', 'data', 'research', 'digital', 'computer', 'engineering', 'mechanical', 'electronic', 'chemical', 'physics', 'biology', 'mathematics', 'information', 'network', 'program', 'algorithm', 'satellite', 'space', 'energy', 'fuel', 'transport', 'vehicle']
        },
        'mind_emotion': {
            'cn': ['心理', '情绪', '情感', '认知', '思考', '记忆', '意识', '感觉', '知觉', '态度', '性格', '行为', '动机', '欲望', '恐惧', '快乐', '悲伤', '愤怒', '焦虑', '压力', '兴奋', '满足', '挫败'],
            'en': ['mental', 'emotion', 'feeling', 'cognitive', 'think', 'memory', 'conscious', 'sense', 'perception', 'attitude', 'personality', 'behavior', 'motivation', 'desire', 'fear', 'happy', 'sad', 'angry', 'anxiety', 'stress', 'excitement', 'satisfaction', 'frustration', 'mind', 'brain']
        },
        'movement_travel': {
            'cn': ['移动', '旅行', '迁移', '流动', '运输', '交通', '行走', '飞行', '航行', '驾驶', '转移', '位置', '方向', '速度', '路径', '路线', '距离'],
            'en': ['move', 'travel', 'migrate', 'wander', 'drift', 'flow', 'shift', 'transfer', 'transport', 'transit', 'motion', 'movement', 'migration', 'flight', 'sail', 'drive', 'walk', 'run', 'fly', 'journey', 'trip', 'voyage', 'route', 'path', 'direction', 'speed']
        },
    }

    for theme, indicators in theme_map.items():
        # Check Chinese definition
        for ind in indicators['cn']:
            if ind in ch:
                return 'theme_cluster'
        # Check keyword
        if kw in indicators['en']:
            return 'theme_cluster'
        # Check replacements
        for r in reps:
            rl = r.lower().strip()
            if rl in indicators['en']:
                return 'theme_cluster'

    # ── Academic expressions ──
    # Long words with Latin/Greek suffixes
    academic_suffixes = ['tion', 'sion', 'sis', 'ical', 'ous', 'ture', 'cally',
                         'ology', 'archy', 'cracy', 'pathy', 'phobia', 'scope',
                         'graph', 'meter', 'metry', 'nomy', 'urgy']
    if len(kw) >= 7:
        for suffix in academic_suffixes:
            if kw.endswith(suffix):
                return 'academic_expression'

    # Words with specific academic prefixes
    academic_prefixes = ['hypo', 'hyper', 'meta', 'para', 'proto', 'pseudo',
                         'neo', 'poly', 'mono', 'multi', 'trans', 'inter',
                         'intra', 'extra', 'counter', 'circum']
    for prefix in academic_prefixes:
        if kw.startswith(prefix) and len(kw) >= 8:
            return 'academic_expression'

    # Known academic words
    academic_keywords_set = {
        'hypothesize', 'postulate', 'extrapolate', 'dichotomy', 'paradigm',
        'empirical', 'qualitative', 'quantitative', 'methodology', 'synthesize',
        'elucidate', 'juxtapose', 'heuristic', 'taxonomy', 'hierarchy',
        'proliferation', 'myriad', 'conundrum', 'preclude', 'ubiquitous',
        'intrinsic', 'extraneous', 'anomaly', 'ephemeral', 'surreptitious',
        'ambivalent', 'didactic', 'pragmatic', 'esoteric', 'pervasive',
        'salient', 'ameliorate', 'exacerbate', 'mitigate', 'disseminate',
        'corroborate', 'substantiate', 'delineate', 'scrutinize', 'perpetuate',
        'necessitate', 'facilitate', 'scrutinize', 'juxtapose',
        'paradoxical', 'counterintuitive', 'disproportionate', 'exponentially',
        'systemic', 'systematic', 'constrain', 'supplement', 'deficiency',
    }
    if kw in academic_keywords_set:
        return 'academic_expression'

    # ── Contextual synonyms ──
    # Words where the replacement is a broader/narrower concept or different context
    # Indicators: Chinese has "n." but keyword is different POS
    # Or the replacement words span multiple semantic fields

    # If replacements are very different in meaning from keyword (based on word roots)
    # this is likely contextual
    if len(reps) >= 3:
        # Check diversity of replacement words
        rep_lens = [len(r) for r in reps]
        if max(rep_lens) - min(rep_lens) > 5:
            return 'contextual_synonym'

    # Mixed POS in replacements suggests contextual
    kw_is_noun = 'n.' in g['chinese'] or kw.endswith('tion') or kw.endswith('sion') or kw.endswith('ment') or kw.endswith('ness') or kw.endswith('ity')
    kw_is_verb = 'v.' in g['chinese'] or kw.endswith('ate') or kw.endswith('ize') or kw.endswith('ify')
    kw_is_adj = 'adj.' in g['chinese'] or kw.endswith('ous') or kw.endswith('ful') or kw.endswith('less') or kw.endswith('ive') or kw.endswith('able') or kw.endswith('ible')

    pos_indicators = 0
    for r in reps:
        rl = r.lower()
        if rl.endswith('ed') or rl.endswith('ing'):
            pos_indicators += 1
        if ' ' in rl:  # phrase
            pos_indicators += 1
        # Check if replacement has different POS from keyword
        if kw_is_noun and (rl.endswith('ed') or rl.endswith('ing') or rl.endswith('ate') or rl.endswith('ize')):
            pos_indicators += 1
        if kw_is_verb and (rl.endswith('tion') or rl.endswith('ment') or rl.endswith('ness') or rl.endswith('ity') or rl.endswith('ous') or rl.endswith('ful')):
            pos_indicators += 1
        if kw_is_adj and (rl.endswith('tion') or rl.endswith('ment') or rl.endswith('ed') or rl.endswith('ing') or rl.endswith('ate')):
            pos_indicators += 1
    if pos_indicators >= 1:
        return 'contextual_synonym'

    # Word boundary shifts: keyword is concrete, replacements are abstract (or vice versa)
    concrete_indicators = ['build', 'make', 'do', 'go', 'come', 'take', 'put', 'get',
                           'see', 'hear', 'eat', 'drink', 'sleep', 'walk', 'run']
    abstract_indicators = ['concept', 'idea', 'notion', 'theory', 'principle', 'phenomenon',
                           'paradigm', 'construct', 'framework', 'approach']

    kw_abstract = any(ai in kw for ai in abstract_indicators)
    reps_abstract = any(any(ai in r.lower() for ai in abstract_indicators) for r in reps)
    if kw_abstract != reps_abstract and len(reps) <= 3:
        return 'contextual_synonym'

    # ── Default: pure_synonym ──
    return 'pure_synonym'


# Re-classify all groups
for g in cleaned_groups:
    g['category'] = classify_group(g)

cat_counts = defaultdict(int)
for g in cleaned_groups:
    cat_counts[g['category']] += 1
print(f'Category distribution: {dict(cat_counts)}')

# Print some logic relations to verify
logic_groups = [g for g in cleaned_groups if g['category'] == 'logic_relation']
print(f'\nLogic relation groups ({len(logic_groups)}):')
for g in logic_groups:
    print(f"  {g['seq']}: {g['keyword']} [{g['chinese']}] → {g['replacements']}")

# ── 4. METADATA ───────────────────────────────────────────────

def assign_theme(g):
    kw = g['keyword'].lower()
    ch = g['chinese']
    reps = [r.lower() for r in g['replacements']]
    all_text = f"{kw} {ch} {' '.join(reps)}"

    theme_map = {
        'nature_environment': ['自然', '环境', '生态', '动物', '植物', '海洋', '气候', '地理', '森林', '物种', '栖息', '天气', '地球', '水', '土', '火', '空气', '太阳', '月', '星', '矿物', '地震', '火山', '洪水', '干旱', '风暴', '灾难', 'earth', 'ocean', 'marine', 'forest', 'climate', 'weather', 'environment', 'soil', 'species', 'habitat', 'wildlife', 'plant', 'animal', 'river', 'mountain', 'water', 'air', 'fire', 'sun', 'moon', 'star', 'flood', 'drought', 'storm', 'earthquake', 'natural', 'wild'],
        'science_technology': ['科学', '技术', '实验', '研究', '数据', '数字', '计算', '工程', '机械', '电子', '化学', '物理', '生物', '数学', '信息', '网络', '电脑', '程序', '算法', '卫星', '太空', '能源', '燃料', 'science', 'technology', 'experiment', 'data', 'research', 'digital', 'computer', 'chemical', 'physics', 'biology', 'energy'],
        'society_culture': ['社会', '文化', '传统', '家庭', '教育', '语言', '政治', '法律', '经济', '政府', '国家', '城市', '农村', '人口', '移民', '阶级', '性别', '种族', '宗教', '历史', '考古', '艺术', '音乐', '建筑', '文学', 'social', 'culture', 'tradition', 'family', 'education', 'language', 'political', 'legal', 'economic', 'government', 'city', 'urban', 'rural', 'population', 'migration', 'history', 'art', 'music', 'architecture'],
        'human_behavior': ['行为', '心理', '认知', '情感', '思考', '记忆', '行为', '意识', '感觉', '知觉', '态度', '性格', '动机', '欲望', '恐惧', '快乐', '悲伤', '愤怒', 'human', 'behavior', 'psychology', 'cognitive', 'emotion', 'memory', 'think', 'conscious', 'sense', 'perception', 'attitude', 'personality', 'motivation', 'fear', 'happy', 'sad', 'angry'],
        'time_change': ['时间', '变化', '发展', '过程', '阶段', '频率', '古代', '现代', '永久', '暂时', '长期', '短期', '延迟', '提前', '快', '慢', 'time', 'change', 'develop', 'process', 'stage', 'frequency', 'ancient', 'modern', 'permanent', 'temporary', 'chronic', 'rapid', 'slow'],
        'quantity_measure': ['数量', '测量', '大小', '比例', '统计', '计算', '巨大', '微小', '多', '少', '尺寸', '规模', '重量', 'quantity', 'measure', 'size', 'proportion', 'massive', 'tiny', 'huge', 'large', 'small', 'statistics', 'scale', 'weight'],
        'health_medicine': ['健康', '疾病', '医疗', '治疗', '药物', '免疫', '病毒', '感染', '症状', '手术', '死亡', 'health', 'disease', 'medical', 'treatment', 'drug', 'immune', 'virus', 'infection', 'symptom', 'surgery', 'death', 'patient', 'doctor'],
        'abstract_concept': ['概念', '理论', '抽象', '哲学', '逻辑', '原理', '现象', '本质', '存在', '意义', '价值', '真理', 'concept', 'theory', 'abstract', 'philosophy', 'logic', 'principle', 'phenomenon', 'essence', 'existence', 'meaning', 'value', 'truth'],
    }

    for theme, indicators in theme_map.items():
        for ind in indicators:
            if ind in all_text:
                return theme

    return 'general_vocabulary'


def calc_difficulty(g):
    kw = g['keyword'].lower()
    reps = g['replacements']
    cat = g['category']

    score = 1

    # Word length contributes
    if len(kw) > 6:
        score += 0.5
    if len(kw) > 8:
        score += 0.5
    if len(kw) > 10:
        score += 0.5

    # Number of replacement words
    if len(reps) > 2:
        score += 0.5
    if len(reps) > 4:
        score += 0.5
    if len(reps) > 6:
        score += 0.5

    # Rare/irregular words
    rare_suffixes = ['tion', 'sion', 'sis', 'ical', 'ous', 'ture', 'ology', 'cally']
    if any(kw.endswith(s) for s in rare_suffixes):
        score += 0.5

    # Category-based adjustment
    cat_modifier = {
        'pure_synonym': 0,
        'theme_cluster': 0.5,
        'contextual_synonym': 0.5,
        'academic_expression': 1,
        'logic_relation': 1.5,
    }
    score += cat_modifier.get(cat, 0)

    # Semantic distance: if replacement words are very different from keyword
    if len(reps) >= 2:
        avg_rep_len = sum(len(r) for r in reps) / len(reps)
        if abs(len(kw) - avg_rep_len) > 5:
            score += 0.5

    # Phrases as replacements (harder to connect)
    phrase_count = sum(1 for r in reps if ' ' in r)
    if phrase_count >= 1:
        score += 0.5
    if phrase_count >= 3:
        score += 0.5

    return min(5, max(1, round(score)))


def calc_semantic_clarity(g):
    kw = g['keyword'].lower()
    reps = g['replacements']
    cat = g['category']

    # Base clarity by category
    cat_base = {
        'pure_synonym': 5,
        'theme_cluster': 4,
        'contextual_synonym': 3,
        'academic_expression': 2,
        'logic_relation': 2,
    }
    base = cat_base.get(cat, 3)

    # More words = harder to see connection
    if len(reps) <= 1:
        base = min(5, base + 1)
    elif len(reps) <= 2:
        base = min(5, base)
    elif len(reps) <= 4:
        base = max(1, base - 1)
    else:
        base = max(1, base - 2)

    # If keyword and replacement share a common root/suffix, easier
    for r in reps:
        rl = r.lower().strip()
        # Check shared suffixes
        for suffix in ['ate', 'ize', 'ify', 'tion', 'ing', 'ed', 'er', 'est', 'ly']:
            if kw.endswith(suffix) and rl.endswith(suffix):
                base = min(5, base + 1)
                break

    # Logic relations are inherently harder to intuit visually
    if cat == 'logic_relation':
        base = max(1, base)

    return base


def determine_gameplay_type(g):
    cat = g['category']
    reps_count = len(g['replacements'])
    total_words = 1 + reps_count

    if cat == 'logic_relation':
        return 'connector_chain'
    elif cat == 'theme_cluster':
        return 'cluster_web' if reps_count >= 2 else 'direct_pair'
    elif total_words <= 2:
        return 'direct_pair'
    elif total_words == 3:
        return 'semantic_chain'
    elif total_words <= 4:
        return 'semantic_chain' if reps_count <= 3 else 'radial_cluster'
    else:
        return 'radial_cluster'


for g in cleaned_groups:
    g['theme'] = assign_theme(g)
    g['difficulty'] = calc_difficulty(g)
    g['semanticClarity'] = calc_semantic_clarity(g)
    g['gameplayType'] = determine_gameplay_type(g)

# Add IDs
for i, g in enumerate(cleaned_groups):
    g['id'] = f'sg_{i+1:04d}'

# ── 5. BUILD FINAL OUTPUT ─────────────────────────────────────

semantic_groups = []
for g in cleaned_groups:
    sg = {
        'id': g['id'],
        'category': g['category'],
        'gameplayType': g['gameplayType'],
        'difficulty': g['difficulty'],
        'semanticClarity': g['semanticClarity'],
        'theme': g['theme'],
        'keyword': g['keyword'],
        'keywordsChinese': g['chinese'],
        'words': [g['keyword']] + g['replacements']
    }
    semantic_groups.append(sg)

# ── Difficulty Progression (redistributed) ────────────────────
# Sort by difficulty then semantic clarity
sorted_by_diff = sorted(semantic_groups, key=lambda g: (g['difficulty'], -g['semanticClarity']))

stages_config = [
    {
        'id': 'stage_1_glow',
        'name': 'Glow — 微光初现',
        'description': '直觉型纯同义词与主题词，光球颜色温暖，语义连接如呼吸般自然',
        'difficulty_range': [1],
        'semanticClarity_min': 4,
        'visual_mood': 'warm_amber',
        'bg_tone': 'golden_hour',
    },
    {
        'id': 'stage_2_tide',
        'name': 'Tide — 潮汐涌动',
        'description': '加入语境型替换，光球开始呈现不同深浅的蓝绿色',
        'difficulty_range': [1, 2],
        'semanticClarity_min': 3,
        'visual_mood': 'ocean_teal',
        'bg_tone': 'twilight_blue',
    },
    {
        'id': 'stage_3_current',
        'name': 'Current — 语义暗流',
        'description': '学术表达与中等难度混合，光球运动轨迹更自由',
        'difficulty_range': [2, 3],
        'semanticClarity_min': 2,
        'visual_mood': 'deep_violet',
        'bg_tone': 'storm_grey',
    },
    {
        'id': 'stage_4_beacon',
        'name': 'Beacon — 灯塔指引',
        'description': '逻辑关系词 + 高难度学术替换，灯塔光束照亮语义路径',
        'difficulty_range': [2, 3, 4],
        'semanticClarity_min': 1,
        'visual_mood': 'lighthouse_gold',
        'bg_tone': 'night_sky',
    },
    {
        'id': 'stage_5_horizon',
        'name': 'Horizon — 天际线',
        'description': '所有类型终极混合，完整语义宇宙',
        'difficulty_range': [3, 4, 5],
        'semanticClarity_min': 1,
        'visual_mood': 'aurora_spectrum',
        'bg_tone': 'dawn_rose',
    },
]

difficulty_progression = {}
for stage in stages_config:
    sid = stage['id']
    difficulty_progression[sid] = {
        'name': stage['name'],
        'description': stage['description'],
        'difficulty_range': stage['difficulty_range'],
        'semanticClarity_min': stage['semanticClarity_min'],
        'visual_mood': stage['visual_mood'],
        'bg_tone': stage['bg_tone'],
        'group_count': 0,
        'group_ids': []
    }

for g in sorted_by_diff:
    d = g['difficulty']
    sc = g['semanticClarity']

    # Assign to appropriate stage with balanced distribution
    # Use a scoring system that spreads groups more evenly
    # stage_score = difficulty * (6 - clarity) — higher = harder
    stage_score = d * (6 - sc)

    if stage_score <= 1:
        # Only purest synonyms (d=1, sc=5)
        difficulty_progression['stage_1_glow']['group_ids'].append(g['id'])
        difficulty_progression['stage_1_glow']['group_count'] += 1
    elif stage_score <= 3:
        # Easy: simple synonyms and themes
        difficulty_progression['stage_2_tide']['group_ids'].append(g['id'])
        difficulty_progression['stage_2_tide']['group_count'] += 1
    elif stage_score <= 6:
        # Medium: academic expressions, moderate clarity
        difficulty_progression['stage_3_current']['group_ids'].append(g['id'])
        difficulty_progression['stage_3_current']['group_count'] += 1
    elif stage_score <= 11:
        # Hard: complex connections
        difficulty_progression['stage_4_beacon']['group_ids'].append(g['id'])
        difficulty_progression['stage_4_beacon']['group_count'] += 1
    else:
        # Zenith: most challenging (score 12+)
        difficulty_progression['stage_5_horizon']['group_ids'].append(g['id'])
        difficulty_progression['stage_5_horizon']['group_count'] += 1

# ── Level Structure ───────────────────────────────────────────
level_structure = []

stage_level_config = [
    ('stage_1_glow', 4, '微光'),
    ('stage_2_tide', 4, '潮汐'),
    ('stage_3_current', 4, '暗流'),
    ('stage_4_beacon', 3, '灯塔'),
    ('stage_5_horizon', 2, '天际'),
]

level_id = 1
for stage_key, num_levels, level_prefix in stage_level_config:
    stage = difficulty_progression[stage_key]
    group_ids = stage['group_ids']

    if not group_ids:
        continue

    groups_per_level = max(1, len(group_ids) // num_levels)

    for i in range(num_levels):
        start = i * groups_per_level
        if i == num_levels - 1:
            end = len(group_ids)
        else:
            end = min(start + groups_per_level, len(group_ids))

        if start >= len(group_ids):
            break

        level_gids = group_ids[start:end]

        # Calculate level-specific stats
        level_groups = [g for g in semantic_groups if g['id'] in level_gids]
        avg_diff = round(sum(g['difficulty'] for g in level_groups) / max(1, len(level_groups)), 1)
        total_word_count = sum(len(g['words']) for g in level_groups)

        level_structure.append({
            'levelId': f'level_{level_id:02d}',
            'stageId': stage_key,
            'stageName': stage['name'],
            'levelName': f'{level_prefix} {i+1}',
            'description': stage['description'],
            'visualMood': stage['visual_mood'],
            'bgTone': stage['bg_tone'],
            'difficultyRange': f'{stage["difficulty_range"]}',
            'groupCount': len(level_gids),
            'totalWordExposures': total_word_count,
            'avgDifficulty': avg_diff,
            'groupIds': level_gids
        })
        level_id += 1

# ── Chain Recommendations ─────────────────────────────────────
chain_recommendations = {
    '2_chain_direct_pair': {
        'name': '双星连线',
        'description': '两个语义最接近的词，浮空光球轻轻碰撞即连接。适合入门，建立"语义连接"直觉。',
        'criteria': 'semanticClarity >= 4, difficulty <= 2',
        'visual': '两个漂浮的暖色光球，碰撞时发出温柔的光芒',
        'group_ids': []
    },
    '3_chain_semantic_flow': {
        'name': '三星语义流',
        'description': '三个词形成线性语义流，渐变光球依次连接。Rescue Duck 核心体验。',
        'criteria': 'semanticClarity >= 3, difficulty <= 3',
        'visual': '三个光球排成弧线，逐一连接如划星星，产生能量流动粒子',
        'group_ids': []
    },
    '4_chain_constellation': {
        'name': '语义星座',
        'description': '四个或更多词形成完整语义星座。适合高级玩家。',
        'criteria': 'semanticClarity >= 2, difficulty <= 4',
        'visual': '多个光球形成星座图案，连接完成时爆发星尘粒子',
        'group_ids': []
    },
}

for g in semantic_groups:
    sc = g['semanticClarity']
    d = g['difficulty']
    word_count = len(g['words'])

    if sc >= 4 and d <= 2:
        chain_recommendations['2_chain_direct_pair']['group_ids'].append(g['id'])
    if sc >= 3 and d <= 3:
        chain_recommendations['3_chain_semantic_flow']['group_ids'].append(g['id'])
    if sc >= 2 and d <= 4:
        chain_recommendations['4_chain_constellation']['group_ids'].append(g['id'])

# ── Late Unlock Recommendations ───────────────────────────────
late_unlock = {
    'description': '后期解锁的 group。需要玩家积累语义直觉后才能驾驭。',
    'subgroups': {
        'high_difficulty': {
            'reason': '高难度组 (difficulty >= 4)，语义连接需要高级直觉',
            'groups': []
        },
        'logic_connectors': {
            'reason': '逻辑关系词，需要建立在大量语义暴露基础上的抽象理解',
            'groups': []
        },
        'low_clarity': {
            'reason': '低语义清晰度组 (semanticClarity <= 2)，连接关系不明显',
            'groups': []
        },
        'academic_mastery': {
            'reason': '高级学术表达，适合学术场景玩家深度挑战',
            'groups': []
        },
    }
}

for g in semantic_groups:
    entry = {'id': g['id'], 'keyword': g['keyword'], 'chinese': g['keywordsChinese']}
    if g['difficulty'] >= 4:
        late_unlock['subgroups']['high_difficulty']['groups'].append(entry)
    if g['category'] == 'logic_relation':
        late_unlock['subgroups']['logic_connectors']['groups'].append(entry)
    if g['semanticClarity'] <= 2:
        late_unlock['subgroups']['low_clarity']['groups'].append(entry)
    if g['difficulty'] >= 3 and g['category'] == 'academic_expression':
        late_unlock['subgroups']['academic_mastery']['groups'].append(entry)

# ── Category Stats ────────────────────────────────────────────
category_stats = defaultdict(lambda: {'count': 0, 'avg_difficulty': 0, 'avg_clarity': 0, 'example_keywords': []})
for g in semantic_groups:
    cat = g['category']
    category_stats[cat]['count'] += 1
    category_stats[cat]['avg_difficulty'] += g['difficulty']
    category_stats[cat]['avg_clarity'] += g['semanticClarity']
    if len(category_stats[cat]['example_keywords']) < 5:
        category_stats[cat]['example_keywords'].append(g['keyword'])

for cat in category_stats:
    n = category_stats[cat]['count']
    category_stats[cat]['avg_difficulty'] = round(category_stats[cat]['avg_difficulty'] / n, 2)
    category_stats[cat]['avg_clarity'] = round(category_stats[cat]['avg_clarity'] / n, 2)

# Overall stats
diff_dist = defaultdict(int)
for g in semantic_groups:
    diff_dist[f'difficulty_{g["difficulty"]}'] += 1

stats = {
    'total_groups': len(semantic_groups),
    'total_unique_words': sum(len(g['words']) for g in semantic_groups),
    'category_distribution': {k: v['count'] for k, v in category_stats.items()},
    'category_details': {k: {'count': v['count'], 'avg_difficulty': v['avg_difficulty'],
                              'avg_clarity': v['avg_clarity'], 'examples': v['example_keywords']}
                         for k, v in category_stats.items()},
    'difficulty_distribution': dict(diff_dist),
    'avg_difficulty': round(sum(g['difficulty'] for g in semantic_groups) / len(semantic_groups), 2),
    'avg_semanticClarity': round(sum(g['semanticClarity'] for g in semantic_groups) / len(semantic_groups), 2),
}

# ── 6. OUTPUT ─────────────────────────────────────────────────
output = {
    'meta': {
        'title': 'Rescue Ducks — Semantic Groups Word Bank',
        'subtitle': '为漂浮光球语义连线体验重构的 IELTS 同义替换词库',
        'version': '2.0.0',
        'designPhilosophy': 'semantic gameplay first — 不是字典式同义词，而是语义能量流',
        'totalGroups': len(semantic_groups),
        'designNotes': [
            '玩家通过反复 semantic exposure 形成长期记忆，而非死记硬背',
            '每个 group 是一个"语义光球簇"，连接难度 = difficulty，直觉度 = semanticClarity',
            '分类优先考虑 gameplay feel 而非语言学严格性',
            'theme 字段用于 atmosphere 场景分组和视觉主题匹配',
        ]
    },
    'stats': stats,
    'semanticGroups': semantic_groups,
    'difficultyProgression': difficulty_progression,
    'levelStructure': level_structure,
    'chainRecommendations': chain_recommendations,
    'lateUnlock': late_unlock,
}

with open('semantic_groups.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

# ── Summary ───────────────────────────────────────────────────
print(f'\n{"="*60}')
print(f'RESCUE DUCKS — SEMANTIC GROUPS GENERATED')
print(f'{"="*60}')
print(f'Total groups: {len(semantic_groups)}')
print(f'Total word exposures: {stats["total_unique_words"]}')
print(f'Avg difficulty: {stats["avg_difficulty"]}')
print(f'Avg semantic clarity: {stats["avg_semanticClarity"]}')
print(f'\nCategory Distribution:')
for cat, count in sorted(stats['category_distribution'].items()):
    print(f'  {cat}: {count} groups')
print(f'\nDifficulty Distribution:')
for d in sorted(diff_dist.keys()):
    print(f'  {d}: {diff_dist[d]} groups')
print(f'\nStages:')
for sid, sdata in difficulty_progression.items():
    print(f'  {sdata["name"]}: {sdata["group_count"]} groups')
print(f'\nLevels: {len(level_structure)}')
print(f'2-chain groups: {len(chain_recommendations["2_chain_direct_pair"]["group_ids"])}')
print(f'3-chain groups: {len(chain_recommendations["3_chain_semantic_flow"]["group_ids"])}')
print(f'4-chain groups: {len(chain_recommendations["4_chain_constellation"]["group_ids"])}')
print(f'Late unlock groups: {sum(len(v["groups"]) for v in late_unlock["subgroups"].values())}')
print(f'\n✓ semantic_groups.json written successfully!')
